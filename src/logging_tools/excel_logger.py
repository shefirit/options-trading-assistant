"""Logs a trade to a local Excel file (the safe fallback / backup).

This writes to its OWN file (trade_log.xlsx in the project folder), never your
teacher's Hebrew tracker, so that file stays safe. Since the "My trades"
tracker was added it can also read the log back, so open positions work even
without a Google Sheet connection.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from src.engine.models import Trade
from src.logging_tools.row import COLUMNS, build_row

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_LOG = PROJECT_ROOT / "trade_log.xlsx"

__all__ = ["COLUMNS", "append_row", "append_values", "read_rows", "delete_trade",
           "DEFAULT_LOG"]


def _open_or_create(path: Path):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        # Older logs have fewer columns - extend the header so new tracker
        # columns (Trade ID, Event...) get labeled.
        have = ws.max_column if ws.max_row >= 1 else 0
        if ws.max_row >= 1 and have < len(COLUMNS):
            for i in range(have, len(COLUMNS)):
                ws.cell(row=1, column=i + 1, value=COLUMNS[i])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append(COLUMNS)
    return wb, ws


def append_values(row: list[Any], path: str | Path = DEFAULT_LOG) -> Path:
    """Append one already-built row (open or close event) to the workbook."""
    path = Path(path)
    wb, ws = _open_or_create(path)
    ws.append(row)
    wb.save(path)
    return path


def append_row(
    trade: Trade,
    strategy_name: str,
    sizing: dict[str, float],
    passed_sop: bool,
    note: str = "",
    path: str | Path = DEFAULT_LOG,
) -> Path:
    """Build the open-event row for a trade and append it (kept for callers/tests)."""
    return append_values(build_row(trade, strategy_name, sizing, passed_sop, note), path)


def read_rows(path: str | Path = DEFAULT_LOG) -> tuple[list[str], list[list[Any]]]:
    """(header, data rows) from the local log. Empty lists if there is no file yet."""
    path = Path(path)
    if not path.exists():
        return [], []
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return [], []
    header = [str(c) if c is not None else "" for c in rows[0]]
    return header, rows[1:]


def delete_trade(trade_id: str, path: str | Path = DEFAULT_LOG) -> int:
    """Remove every row for one trade (by Trade ID) from the local log.
    Returns how many rows were removed."""
    path = Path(path)
    if not path.exists() or not trade_id:
        return 0
    wb = load_workbook(path)
    ws = wb.active
    try:
        id_col = COLUMNS.index("Trade ID") + 1   # openpyxl is 1-based
    except ValueError:
        return 0
    to_delete = [r for r in range(2, ws.max_row + 1)
                 if str(ws.cell(row=r, column=id_col).value or "") == str(trade_id)]
    for r in reversed(to_delete):   # bottom-up so row numbers stay valid
        ws.delete_rows(r, 1)
    if to_delete:
        wb.save(path)
    return len(to_delete)
