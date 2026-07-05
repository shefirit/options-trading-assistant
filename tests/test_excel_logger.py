"""Test the trade logger writes rows correctly (into a temp file)."""

from __future__ import annotations

from openpyxl import load_workbook

from src.engine.models import Action, Leg, OptionType, Trade
from src.logging_tools.excel_logger import COLUMNS, append_row


def _trade() -> Trade:
    return Trade(
        strategy_key="put_credit_spread", underlying="SPX", contracts=3,
        legs=[
            Leg(role="short_put", action=Action.SELL, option_type=OptionType.PUT,
                strike=5000, delta=-0.08, premium=8.0, dte=45),
            Leg(role="long_put", action=Action.BUY, option_type=OptionType.PUT,
                strike=4975, delta=-0.05, premium=5.0, dte=45),
        ],
    )


def test_creates_file_with_header_and_row(tmp_path):
    path = tmp_path / "log.xlsx"
    size = {"credit": 900.0, "max_loss": 6600.0, "buying_power": 6600.0}
    append_row(_trade(), "Put Credit Spread", size, passed_sop=True, note="test", path=path)

    wb = load_workbook(path)
    ws = wb.active
    assert [c.value for c in ws[1]] == COLUMNS
    row = [c.value for c in ws[2]]
    assert "SPX" in row
    assert "Put Credit Spread" in row
    assert "yes" in row  # passed_sop


def test_appends_second_row(tmp_path):
    path = tmp_path / "log.xlsx"
    size = {"credit": 900.0, "max_loss": 6600.0, "buying_power": 6600.0}
    append_row(_trade(), "Put Credit Spread", size, True, "one", path=path)
    append_row(_trade(), "Put Credit Spread", size, False, "two", path=path)

    wb = load_workbook(path)
    ws = wb.active
    assert ws.max_row == 3  # header + 2 rows
    assert [c.value for c in ws[3]][11] == "two"   # the Notes column
    assert "NO" in [c.value for c in ws[3]]  # failed_sop marked


def test_read_rows_round_trip(tmp_path):
    from src.logging_tools.excel_logger import read_rows
    path = tmp_path / "log.xlsx"
    size = {"credit": 900.0, "max_loss": 6600.0, "buying_power": 6600.0}
    append_row(_trade(), "Put Credit Spread", size, True, "hello", path=path)
    header, rows = read_rows(path)
    assert header == COLUMNS
    assert len(rows) == 1
    assert rows[0][1] == "SPX"


def test_read_rows_missing_file(tmp_path):
    from src.logging_tools.excel_logger import read_rows
    header, rows = read_rows(tmp_path / "nope.xlsx")
    assert header == [] and rows == []


def test_old_log_header_gets_extended(tmp_path):
    """A log created before the tracker existed gains the new column labels."""
    from openpyxl import Workbook
    from src.logging_tools.excel_logger import append_values, read_rows
    path = tmp_path / "old.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(COLUMNS[:12])                      # the original 12 columns
    ws.append(["2026-06-20", "SPX", "Put Credit Spread", "5000 / 4975",
               0.08, 45, 1, 300, 2200, 2200, "yes", "legacy trade"])
    wb.save(path)

    append_values(["2026-07-05", "SPX", "Put Credit Spread", "", "", "", "",
                   "", "", "", "", "closed it", "20260705-1-SPX", "close", "",
                   150.0, 150.0, ""], path)
    header, rows = read_rows(path)
    assert header == COLUMNS
    assert len(rows) == 2
